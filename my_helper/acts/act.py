"""
акты в xlsx
Открываем пустой акт и заполняем
Потом программа формирует ведомость материалов
Указываем коэффициенты по материалам для пересчета
шт только целое число можжет быть
*может делать свой вариант ведомости, преобразовывать ее, что бы было проще потом работать
Акт готов
__
Создаем новый акт за месяц
в форме указываем номер пункта.
Программ заполняет поля из табл и пишет, сколько осталось по нему
редактируем и нажимаем добавить
Так добавляем все пункты
Потом открываем получившийся акт
Открываем окно для табеля дней с запоненными людьми
Указываем дни
Формируется выработка

"""
import openpyxl as pyxls

name = 0
point = 1
text = 2
SI = 3
count = 4
price = 5
all_fields = [text, SI, count, price]
self_material = "Материалы Подрядчика"
fabric_material = "Материалы Заказчика"

def get_val(sheet, row, column):
    if column is int:
        cell = sheet.cell(row=row, column=column)
        return cell.value
    if column is list:
        vals = list()
        for i in range(10):
            cell = sheet.cell(row=row, column=column[i])
            vals.append(cell.value)
        return vals

def set_val(sheet, row, column, val):
    if val is int:
        cell = sheet.cell(row=row, column=column)
        cell.value = val
        return
    if column is list:
        i = iter(range(10))
        for item in val:
            cell = sheet.cell(row=row, column=column + next(i))
            cell.value = item
            return


class Acts:
    def create_material_doc(self, wb):
        """
        создать ведомость материалов
        :return:
        """
        xls = pyxls.load_workbook(wb)
        sh = xls["act"]
        check = 0
        _my_mat = list()
        _not_my_mat = list()
        i = iter(range(1000))
        while check < 3:
            next(i)
            value = get_val(sh, i, text)
            check += 1 if value == "" else check
            if value == self_material:
                while True:
                    next(i)
                    value = get_val(sh, i, point)
                    if value == "":
                        _my_mat.append([0, get_val(sh, i, all_fields)])  # добавляем материал
            elif value == fabric_material:
                while True:
                    next(i)
                    value = get_val(sh, i, point)
                    if value == "":
                        _not_my_mat.append([0, get_val(sh, i, all_fields)])
        my_mat = set(_my_mat)
        not_my_mat = set(_not_my_mat)
        for mat in my_mat:
            for item in _my_mat:
                if item[name] == mat[name]:
                    mat[count] = mat[count] + item[count]
        for mat in not_my_mat:
            for item in _not_my_mat:
                if item[name] == mat[name]:
                    mat[count] = mat[count] + item[count]
        sh = xls["my_mat"]
        i = iter(range(1000))
        for item in my_mat:
            set_val(sh, next(i), 0, item)



    def calc_all_materials(self):
        """
        подсчет всех материалов по ведомости
        :return:
        """
        pass

    def calc_price_for_worker(self):
        """
        подсчет для выроботки
        :return:
        """
        pass

    def calc_price_act(self):
        """
        подсчет суммы по акту
        :return:
        """
        pass

    def create_report_materials(self):
        """
        отчет по материалам
        Всего по акту, осталось по акту, получили, израсходывали, осталось.
        :return:
        """

    def create_report_act(self):
        """
        Отчет по акту. Какие пункты остались по акту
        :return:
        """
        pass
