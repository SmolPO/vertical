from PyQt5.QtWidgets import QMessageBox as mes
from PyQt5.QtWidgets import QLabel, QCheckBox, QSpinBox, QDialog
import openpyxl as xlsx
import os
from PyQt5 import uic
import datetime as dt
designer_file = '../designer_ui/tabel.ui'


class NewTable(QDialog):
    def __init__(self, parent=None):
        super(NewTable, self).__init__(parent)
        uic.loadUi(designer_file, self)
        self.parent = parent
        self.table = "workers"
        self.count = self.parent.count_people_tb
        self.b_ok.clicked.connect(self.ev_ok)
        self.b_cancel.clicked.connect(self.ev_cancel)
        self.rows_from_db = self.parent.db.get_data("*", self.table)
        self.table_file = ""
        self.init_list()

    def init_list(self):
        g = iter(range(1, self.count + 1))
        for item in self.rows_from_db:
            i = next(g)
            number = QLabel(item[-1])
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.ev_select)
            family = QLabel(item[0] + item[1][0] + "." + item[2][0] + ".")
            time = QSpinBox()
            time.setValue(8)
            self.grid.addWidget(number, i, 0)
            self.grid.addWidget(checkbox, i, 1)
            self.grid.addWidget(family, i, 2)
            self.grid.addWidget(time, i, 3)

    def get_list_people(self): # _type = 1 or 2
        list_people = list()
        for i in range(self.count):
            if self.grid.itemAtPosition(i, 1).widget().isChecked():
                number = self.grid.itemAtPosition(i, 0).widget().text()
                time = self.grid.itemAtPosition(i, 3).widget().value()
                list_people.append([self.rows_from_db[int(number)-1], time])
        return list_people

    def ev_ok(self):
        workers = self.get_list_people()
        self.create_table(workers)
        self.close()
        return True

    def create_table(self, people):
        doc = xlsx.open("/covid.xlsx")
        sheet = doc["covid"]
        print(people)
        for i in range(len(people)):
            sheet.cell(row=i + 2, column=1).value = people[-1]
            sheet.cell(row=i + 2, column=2).value = str(dt.datetime.now().date())
            sheet.cell(row=i + 2, column=3).value = people[2] + " " + people[3][0] + "."
            sheet.cell(row=i + 2, column=4).value = people[4]
        file_path = "/covid/" + str(dt.datetime.now().date()) + "_" + self.parent.company
        doc.save(file_path)
        os.startfile(file_path, "print")

        list_people = list()
        for i in range(self.count):
            time = self.grid.itemAtPosition(i, 3).widget().value()
            list_people.append(time)

        doc = xlsx.open("/table.xlsx")
        sheet = doc[str(dt.datetime.now().month)]
        count = 0
        for i in range(len(self.rows_from_db)):
            if sheet.cell(row=i + 2, column=1).value:
                count = sheet.cell(row=i + 2, column=1).value
        if count != len(self.rows_from_db):
        for i in range(len(self.rows_from_db)):
            val = "н" if list_people[i + 2] == 0 else list_people[i + 2]
            sheet.cell(row=i + 2, column=dt.datetime.now().day + 3).value = val
        doc.save("/table.xlsx")
        return

    def ev_select(self, state):
        for i in range(self.count):
            if self.grid.itemAtPosition(i, 1).widget().isChecked():
                self.grid.itemAtPosition(i, 0).widget().setEnabled(state)
                self.grid.itemAtPosition(i, 2).widget().setEnabled(state)
                self.grid.itemAtPosition(i, 3).widget().setEnabled(state)
                self.grid.itemAtPosition(i, 3).widget().setValue(0)
