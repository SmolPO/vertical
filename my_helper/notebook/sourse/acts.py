import openpyxl as pyxls


class CreateActs:
    def create_list_materils(self, xlsx):
        # откыть док
        wb = pyxls.load_workbook(xlsx)
        sheet = wb["acts"]
        for i in range(1, 4):
            value_1 = sheet.cell(row=i, column=1).value
            value_2 = sheet.cell(row=i, column=2).value
            if int(value_1):
                pass

        # читаем и находим все позиции
        # если такой позиции нет в отчете - добавляем и суммируем объем
        pass

    def get_price_for_production(self):
        # открыть док
        # находим слово Работа или там где нет ничего, кроме позиции
        # суммируем стоимость
        pass
