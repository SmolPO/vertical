import openpyxl as xlsx
from PyQt5.QtWidgets import QMessageBox as mes
from openpyxl.styles import Side, Border
import os
import datetime as dt
from PyQt5.QtWidgets import QInputDialog
from my_helper.notebook.sourse.database import *
month = ["январь", "февраль", "март", "апрель", "май", "июнь",
         "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]


class NewTable:
    def __init__(self, parent=None):
        self.parent = parent
        self.conf = Ini(self)
        self.border = Border(top=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             left=Side(border_style='thin', color='FF000000'))

    def create_table(self):
        list_people = list()
        try:
            rows = self.parent.db.get_data("family, name, surname, status, id", "workers") + \
                   self.parent.db.get_data("family, name, surname, status, id", "itrs")
        except:
            msg_er(self, GET_DB)
            return
        for row in rows:
            if row[-2] != statues[2]:
                list_people.append(short_name(row))
        list_people.sort()
        cur_month, ok = QInputDialog.getItem(self.parent, "Выберите месяц", "Месяц", month, 0)
        if not ok:
            return
        try:
            path = self.conf.get_path("path") + self.conf.get_path("path_pat_tabel")
        except:
            msg_er(self, GET_INI)
            return
        path_save = path
        try:
            doc = xlsx.open(path)
        except:
            msg_er(self, GET_FILE + path)
            return
        try:
            page = "табель_" + str(count_days[month.index(cur_month)])
            sheet = doc[page]
        except:
            msg_er(self, GET_PAGE + page)
            return
        delta = 5
        count_column = count_days[month.index(cur_month)] + 2
        sheet.cell(row=2, column=1).value = str(cur_month) + " " + str(dt.datetime.now().year)
        for ind in range(1, 50):
            for j in range(1, count_column):
                sheet.cell(row=ind + delta, column=j).value = ""
        for ind in range(1, len(list_people) + 5):
            sheet.cell(row=ind + delta, column=1).value = ind
            sheet.cell(row=ind + delta, column=2).value = list_people[ind - 1]
            for i in range(1, count_column):
                sheet.cell(row=ind + delta, column=i).border = self.border
        doc.print_area = "A1:AF" + str(len(list_people) + delta + 5)
        try:
            doc.save(path_save)
            os.startfile(path_save)
        except:
            msg_er(self, GET_FILE + path_save)

